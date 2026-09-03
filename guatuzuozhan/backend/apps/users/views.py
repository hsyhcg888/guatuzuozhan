import logging
import secrets
import string
from io import BytesIO
from openpyxl import Workbook, load_workbook
from django.db import transaction
from django.utils import timezone
from rest_framework import status
from rest_framework.pagination import PageNumberPagination
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_simplejwt.tokens import RefreshToken
from .models import Department, User
from .permissions import IsSystemAdmin
from .serializers import AdminUserWriteSerializer, DepartmentSerializer, LoginSerializer, RegisterSerializer, UserSerializer

logger = logging.getLogger('user_audit')

def token_response(user):
    refresh = RefreshToken.for_user(user)
    return {'access': str(refresh.access_token), 'refresh': str(refresh), 'user': UserSerializer(user).data}

class RegisterView(APIView):
    permission_classes = [AllowAny]
    def post(self, request):
        serializer = RegisterSerializer(data=request.data); serializer.is_valid(raise_exception=True)
        data = serializer.validated_data; data.pop('confirm_password', None); password = data.pop('password')
        user = User.objects.create_user(password=password, account_status='disabled', register_status='pending', is_system_admin=False, is_deleted=False, **data)
        return Response({'message': '注册申请已提交，请等待系统管理员审核', 'id': user.id}, status=status.HTTP_201_CREATED)

class LoginView(APIView):
    permission_classes = [AllowAny]
    def post(self, request):
        s = LoginSerializer(data=request.data); s.is_valid(raise_exception=True)
        user = User.objects.filter(name=s.validated_data['name'], is_deleted=False).first()
        if not user or not user.check_password(s.validated_data['password']): return Response({'detail': '姓名或密码错误'}, status=400)
        if user.register_status == 'pending': return Response({'detail': '您的注册申请正在审核中，请等待系统管理员审核'}, status=403)
        if user.register_status == 'rejected': return Response({'detail': '您的注册申请未通过审核'}, status=403)
        if user.account_status != 'enabled': return Response({'detail': '该账号已被禁用，请联系系统管理员'}, status=403)
        user.last_login_at = timezone.now(); user.save(update_fields=['last_login_at']); return Response(token_response(user))

class MeView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request): return Response(UserSerializer(request.user).data)

class DepartmentListView(APIView):
    permission_classes = [AllowAny]
    def get(self, request): return Response(DepartmentSerializer(Department.objects.filter(is_deleted=False, status=1).order_by('sort_order'), many=True).data)

class AdminUsersView(APIView):
    permission_classes = [IsSystemAdmin]
    def get(self, request):
        qs = User.objects.filter(is_deleted=False).select_related('department').order_by('-created_at')
        for field in ('name', 'position', 'account_status', 'register_status'):
            if request.query_params.get(field): qs = qs.filter(**{field + '__icontains' if field == 'name' else field: request.query_params[field]})
        if request.query_params.get('department_id'): qs = qs.filter(department_id=request.query_params['department_id'])
        paginator = PageNumberPagination(); page = paginator.paginate_queryset(qs, request); return paginator.get_paginated_response(UserSerializer(page, many=True).data)
    def post(self, request):
        s = AdminUserWriteSerializer(data=request.data); s.is_valid(raise_exception=True); data=s.validated_data; password=data.pop('password', None)
        if not password: return Response({'password': '新增用户必须提供密码'}, status=400)
        user=User.objects.create_user(password=password, is_deleted=False, **data); logger.info('admin_create user=%s by=%s', user.id, request.user.id); return Response(UserSerializer(user).data, status=201)

class AdminUserDetailView(APIView):
    permission_classes = [IsSystemAdmin]
    def get_object(self, pk): return User.objects.filter(pk=pk, is_deleted=False).first()
    def get(self, request, pk):
        user=self.get_object(pk); return Response(UserSerializer(user).data) if user else Response(status=404)
    def patch(self, request, pk):
        user=self.get_object(pk)
        if not user: return Response(status=404)
        s=AdminUserWriteSerializer(data={**request.data, 'password': request.data.get('password', 'temporary')}, partial=True); s.is_valid(raise_exception=True)
        data=s.validated_data; password=data.pop('password', None)
        if password: user.set_password(password)
        for k,v in data.items(): setattr(user, k if k != 'department_id' else 'department_id', v)
        user.updated_at=timezone.now(); user.save(); logger.info('admin_update user=%s by=%s', user.id, request.user.id); return Response(UserSerializer(user).data)
    def delete(self, request, pk):
        user=self.get_object(pk)
        if not user: return Response(status=404)
        if user.is_system_admin and User.objects.filter(is_system_admin=True, is_deleted=False).count() <= 1: return Response({'detail':'不能删除最后一个系统管理员'}, status=400)
        user.is_deleted=True; user.deleted_at=timezone.now(); user.save(update_fields=['is_deleted','deleted_at','updated_at']); logger.info('admin_delete user=%s by=%s', user.id, request.user.id); return Response(status=204)

class AdminActionView(APIView):
    permission_classes = [IsSystemAdmin]
    def post(self, request, pk, action):
        user=User.objects.filter(pk=pk, is_deleted=False).first()
        if not user: return Response(status=404)
        if action == 'approve': user.register_status='approved'; user.account_status='enabled'; user.review_remark=request.data.get('remark')
        elif action == 'reject':
            if not request.data.get('remark'): return Response({'remark':'驳回原因必填'}, status=400)
            user.register_status='rejected'; user.account_status='disabled'; user.review_remark=request.data['remark']
        elif action == 'enable': user.account_status='enabled'
        elif action == 'disable':
            if user.is_system_admin and User.objects.filter(is_system_admin=True, is_deleted=False, account_status='enabled').count() <= 1: return Response({'detail':'不能禁用最后一个系统管理员'}, status=400)
            user.account_status='disabled'
        elif action == 'reset-password':
            password = request.data.get('password') or ''.join(secrets.choice(string.ascii_letters + string.digits) for _ in range(12))
            user.set_password(password); user.updated_at=timezone.now(); user.save(update_fields=['password_hash','updated_at'])
            logger.info('admin_reset_password user=%s by=%s', user.id, request.user.id)
            return Response({'message': '密码已重置', 'temporary_password': password})
        else: return Response(status=400)
        user.reviewer_id=request.user.id; user.reviewed_at=timezone.now(); user.updated_at=timezone.now(); user.save(); logger.info('admin_%s user=%s by=%s', action, user.id, request.user.id); return Response(UserSerializer(user).data)

class UserTemplateView(APIView):
    permission_classes = [IsSystemAdmin]
    def get(self, request):
        workbook = Workbook(); sheet = workbook.active
        sheet.append(['姓名', '初始密码', '职位', '部门编码', '用户岗位', '用户职责描述', '是否系统管理员', '账号状态', '审核状态'])
        output = BytesIO(); workbook.save(output); output.seek(0)
        from django.http import HttpResponse
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="users_template.xlsx"'; return response

class UserImportView(APIView):
    permission_classes = [IsSystemAdmin]
    def post(self, request):
        upload = request.FILES.get('file')
        if not upload: return Response({'detail': '请上传 Excel 文件'}, status=400)
        try: rows = list(load_workbook(upload, read_only=True, data_only=True).active.iter_rows(values_only=True))
        except Exception: return Response({'detail': 'Excel 文件格式无效'}, status=400)
        results=[]; headers=rows[0] if rows else []
        expected=['姓名','初始密码','职位','部门编码','用户岗位','用户职责描述','是否系统管理员','账号状态','审核状态']
        if list(headers) != expected: return Response({'detail': 'Excel 表头不正确', 'expected': expected}, status=400)
        names=set()
        for number, row in enumerate(rows[1:], 2):
            data=dict(zip(expected, row)); errors=[]; name=(data.get('姓名') or '').strip()
            if not name: errors.append('姓名不能为空')
            if name in names: errors.append('Excel 内姓名重复')
            if User.objects.filter(name=name).exists(): errors.append('姓名已存在')
            names.add(name); password=data.get('初始密码')
            if not password: errors.append('初始密码不能为空')
            if data.get('职位') not in {'general_manager','deputy_general_manager','section_chief','deputy_section_chief','team_leader','member'}: errors.append('职位编码不合法')
            department=Department.objects.filter(code=data.get('部门编码'), is_deleted=False).first()
            if not department: errors.append('部门编码不存在')
            if errors: results.append({'row': number, 'success': False, 'errors': errors}); continue
            user=User.objects.create_user(name=name, password=str(password), position=data['职位'], department_id=department.id, job_title=data.get('用户岗位') or None, responsibility=data.get('用户职责描述') or None, is_system_admin=bool(data.get('是否系统管理员')), account_status=data.get('账号状态') or 'disabled', register_status=data.get('审核状态') or 'pending', is_deleted=False)
            results.append({'row': number, 'success': True, 'id': user.id})
        return Response({'total': len(results), 'success_count': sum(x['success'] for x in results), 'results': results})
